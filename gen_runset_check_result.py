#!/bin/env python3.7
import os
import re
import sys
from optparse import OptionParser
import csv
import openpyxl

def help():
    print('''
****************************************************************
=> Usage:
    % script -> $0

    % gen_runset_check_result.py --process <process> --metal <metal>

    eg: gen_runset_check_result.py --process cxmt10G5plus --metal 1p5m1x3y1z

    Author: ssfang

****************************************************************
    ''')

def ChangeColumnWidth( sheet, column, width ):

  columnLetter = openpyxl.utils.get_column_letter( column )
  sheet.column_dimensions[ columnLetter ].width = width

def RunsetCheck(deck, CSVresult):
    write = open( CSVresult, "w" )
    csvWriter = csv.writer(write)
    csvWriter.writerow([ "NUM", "ITEM", "STATUS", "REMARK" ])
    num = 0
    for line in open( deck, "r" ):
        onDefine = re.match(r'\s*(#DEFINE\s+[^//]+)(//.*)?', line)
        offDefine = re.match(r'\s*(//#DEFINE\s+[^//]+)(//.*)?', line)
        if onDefine:
            #print( onDefine.group(1), onDefine.group(2) )
            num += 1
            csvWriter.writerow([ num, onDefine.group(1), "OPEN", onDefine.group(2) ])
        elif offDefine:
            #print( offDefine.group(1), offDefine.group(2) )
            num += 1
            csvWriter.writerow([num, offDefine.group(1), "CLOSE", offDefine.group(2) ])
        else:
            continue

    write.close()

def GenResult( processDir ):
    DRCdeck = str( processDir ) + "/calibreDRC.rule"
    DRCRunsetCheck = RunsetCheck( DRCdeck, "DRC_runset_check.csv" )

    LVSdeck = str( processDir ) + "/calibreLVS.rule"
    LVSRunsetCheck = RunsetCheck( LVSdeck, "LVS_runset_check.csv" )

    ERCdeck = str( processDir ) + "/calibreERC.rule"
    ERCRunsetCheck = RunsetCheck( ERCdeck, "ERC_runset_check.csv" )

    ANTdeck = str( processDir ) + "/calibreANT.rule"
    ANTRunsetCheck = RunsetCheck( ANTdeck, "ANT_runset_check.csv" )

    LPEdeck = str( processDir ) + "/lpe_star/calibreLPE.rule"
    LPERunsetCheck = RunsetCheck( LPEdeck, "LPE_runset_check.csv" )

def CreateRunsetCheckExcel( workBook, csvFile ):
    sheet = workBook.create_sheet( re.sub( r'\.csv$', "", str(csvFile) ) )
    read = open( csvFile, "r" )    
    csvRead = csv.reader( read )
    
    row = 1
    column = 1

    headRow = list()
    headRow.append(1)
    for line in csvRead:
        if len(line) > 0:
            for token in line:
                sheet.cell( row, column ).font = gFontCalibri
                sheet.cell( row, column ).value = token
                sheet.cell( row, column ).border = gThinBorder
                column += 1

            sheet.cell( row, 1 ).alignment = gAlignCenter
            sheet.cell( row, 2 ).alignment = gAlignCenter
            sheet.cell( row, 3 ).alignment = gAlignCenter

        else:
            headRow.append( row + 1 )

        row += 1
        column = 1

    read.close()
    
    #Setup sheet head row.
    for row in headRow:
        for column in range( 1, 5 ):
            sheet.cell( row, column ).font = gFontCalibriAndBold
            sheet.cell( row, column ).alignment = gAlignCenter
            sheet.cell( row, column ).fill = gBackgroundColorYellow
    
    #Setup row color and height. 
    for i in range( 1, sheet.max_row+1 ):
        sheet.row_dimensions[i].height = 22
        if i%2 == 1:
            for j in range( 1, sheet.max_column+1 ):
                sheet.cell( row = i, column = j ).fill = gBackgroundColorYellow
  
    ChangeColumnWidth( sheet, 1, 15 )
    ChangeColumnWidth( sheet, 2, 50 )
    ChangeColumnWidth( sheet, 3, 20 )
    ChangeColumnWidth( sheet, 4, 60 )

    
gBackgroundColorYellow = openpyxl.styles.PatternFill( start_color = "FAEBD7", end_color = "FAEBD7", fill_type = "solid" )

gFontCalibri = openpyxl.styles.Font( name = "Calibri" )
gFontCalibriAndBold = openpyxl.styles.Font( name = "Calibri", bold = True )
gThin = openpyxl.styles.borders.Side( border_style = openpyxl.styles.borders.BORDER_THIN )
gAlignCenter = openpyxl.styles.Alignment( horizontal = "center", vertical = "center" )
gThinBorder = openpyxl.styles.borders.Border( left = gThin, right = gThin, top = gThin, bottom = gThin )  

if __name__ == "__main__":
    parse = OptionParser()
    parse.add_option('--process', dest = 'process')
    parse.add_option('--metal', dest = 'metal')
    options = parse.parse_args()[0]
    optProcess = options.process
    optMetal = options.metal

    if (optProcess is None) | (optMetal is None):
        help()
    
    processDir = "/apps/imctf/runset/calibre/" + str( optProcess ) + "/current/" + str( optMetal )

    if not os.path.exists( processDir ):
        print( "Incorrect process and metal name! Please rerun!\n" )

    currentDir = os.getcwd()
    runsetChecks = optProcess + "_Runset_Check"
    if os.path.exists( runsetChecks ):
        os.remove( runsetChecks )
    os.mkdir( runsetChecks )
    os.chdir(currentDir + "/" + runsetChecks)
    GenResult(processDir)
    workBook = openpyxl.Workbook()

    for csvFile in os.listdir( os.getcwd() ):
        if str(csvFile).endswith( "csv" ):
            CreateRunsetCheckExcel( workBook, csvFile )
        else:
            continue

    workBook.remove(workBook[ "Sheet" ])
    workBook.save( str( optProcess ) + "_Runset_Check.xlsx" )

